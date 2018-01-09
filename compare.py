from openpyxl import load_workbook
# The source xlsx file is named as source.xlsx
wb=load_workbook("teine.xlsx")
wb2=load_workbook('KASTINUMBRID.XLSX')

ws = wb.active
ws2= wb2.active
first_column = ws['A']
second_column = ws2['A']

# Print the contents
for x in range(len(first_column)):
    print(first_column[x].value)